import streamlit as st
import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Agency Opportunity Scorecard", page_icon="🎯", layout="centered")

st.markdown("""
<style>
    .verdict-box { border-radius: 12px; padding: 24px; text-align: center; margin-top: 24px; }
    .strong { background-color: #d4edda; border: 2px solid #28a745; }
    .possible { background-color: #fff3cd; border: 2px solid #ffc107; }
    .longshot { background-color: #f8d7da; border: 2px solid #dc3545; }
    .score-big { font-size: 64px; font-weight: 800; line-height: 1; }
    .verdict-label { font-size: 22px; font-weight: 600; margin-top: 8px; }
    .section-header { font-size: 17px; font-weight: 700; color: #1a1a2e; border-left: 4px solid #4361ee; padding-left: 10px; margin-top: 28px; margin-bottom: 8px; }
</style>
""", unsafe_allow_html=True)

st.title("🎯 Agency Opportunity Scorecard")
st.caption("Evaluate whether a new business opportunity is worth pursuing.")

QUESTIONS = {
    "Competitive Landscape": [
        {"id": "Q1",  "label": "Number of competing agencies",       "hint": "Fewer competitors = higher score",           "max": 20,
         "options": [("1–2 agencies", 20), ("3–4 agencies", 6), ("5–6 agencies", 2), ("7+ agencies", 0)]},
        {"id": "Q2",  "label": "Incumbent status",                   "hint": "Are we defending or challenging?",            "max": 8,
         "options": [("We are the incumbent", 8), ("Challenger — incumbent is weak", 5), ("No incumbent / unknown", 4), ("Challenger — incumbent is strong", 2)]},
        {"id": "Q3",  "label": "Do you know who the other agencies are?", "hint": "Known competitors allow better strategy", "max": 4,
         "options": [("Yes — all known & we compare well", 4), ("Yes — but competition is strong", 2), ("No — mostly unknown", 1)]},
    ],
    "Relationship & History": [
        {"id": "Q4",  "label": "Prior relationship with client",      "hint": "Existing relationships increase win rate",    "max": 8,
         "options": [("Previous client — positive history", 8), ("Met them but never worked together", 4), ("No prior relationship", 1)]},
        {"id": "Q5",  "label": "Pitch origin",                       "hint": "Inbound pitches signal stronger intent",      "max": 6,
         "options": [("Inbound — they approached us", 6), ("Outbound — we approached them", 2)]},
        {"id": "Q6",  "label": "Internal champion / sponsor",        "hint": "An advocate inside the brand is a strong signal", "max": 6,
         "options": [("Yes — identified & engaged", 6), ("Possible but unconfirmed", 3), ("No internal champion", 0)]},
        {"id": "Q7",  "label": "LinkedIn connections at the brand",  "hint": "Network depth at the client",                 "max": 4,
         "options": [("5+ connections", 4), ("2–4 connections", 2), ("1 connection", 1), ("None", 0)]},
    ],
    "Brief & Process": [
        {"id": "Q8",  "label": "Brief quality & alignment",          "hint": "How well does it match our strengths?",       "max": 6,
         "options": [("Detailed brief — strong fit", 6), ("Good brief — partial fit", 4), ("Vague brief or weak fit", 1)]},
        {"id": "Q9",  "label": "Access to decision-maker",           "hint": "Chemistry meeting or direct contact",         "max": 4,
         "options": [("Yes — chemistry meeting held", 4), ("Some contact but no formal meeting", 2), ("No access to decision-maker", 0)]},
        {"id": "Q10", "label": "Time given to prepare",              "hint": "More time benefits challengers",              "max": 3,
         "options": [("3+ weeks", 3), ("1–3 weeks", 2), ("Under 1 week", 0)]},
    ],
    "Commercial Opportunity": [
        {"id": "Q11", "label": "Overall $ value of the opportunity", "hint": "Larger accounts justify more investment",     "max": 5,
         "options": [("$500k+", 5), ("$200k–$500k", 4), ("$100k–$200k", 3), ("Under $100k", 1)]},
        {"id": "Q12", "label": "Growth potential beyond initial brief", "hint": "Can this account expand over time?",       "max": 3,
         "options": [("High — clear expansion opportunity", 3), ("Medium — some potential", 2), ("Low — likely stays as scoped", 1)]},
        {"id": "Q13", "label": "Budget realism",                    "hint": "Does the budget match the scope?",             "max": 3,
         "options": [("Budget is realistic for scope", 3), ("Slightly underfunded but workable", 2), ("Budget too low for scope", 0)]},
    ],
    "Scope & Category": [
        {"id": "Q14", "label": "Number of channels pitched for",     "hint": "More channels = more value & stickiness",     "max": 5,
         "options": [("5+ channels", 5), ("3–4 channels", 4), ("2 channels", 2), ("1 channel", 1)]},
        {"id": "Q15", "label": "Client category",                   "hint": "How strong is your category expertise?",       "max": 5,
         "options": [("Core category (fashion / beauty)", 5), ("Adjacent category", 3), ("New/unfamiliar category (automotive)", 1)]},
        {"id": "Q16", "label": "Strategic fit for agency portfolio", "hint": "Does winning help your agency's story?",       "max": 3,
         "options": [("Trophy client or new category win", 3), ("Good fit — not transformational", 2), ("Limited strategic value", 1)]},
    ],
}

MAX_PER_CATEGORY = {
    "Competitive Landscape": 32,
    "Relationship & History": 24,
    "Brief & Process": 13,
    "Commercial Opportunity": 11,
    "Scope & Category": 13,
}

# ── Collect answers ──────────────────────────────────────────────────────────
answers = {}
total_questions = sum(len(qs) for qs in QUESTIONS.values())

for category, questions in QUESTIONS.items():
    max_pts = MAX_PER_CATEGORY[category]
    st.markdown(f'<div class="section-header">{category} &nbsp;<span style="font-weight:400;font-size:13px;color:#888">Max {max_pts} pts</span></div>', unsafe_allow_html=True)
    for q in questions:
        option_labels = [o[0] for o in q["options"]]
        choice = st.selectbox(
            f"{q['label']} — *{q['hint']}*",
            options=["— Select —"] + option_labels,
            key=q["id"],
        )
        answers[q["id"]] = dict(q["options"]).get(choice)

st.divider()

answered = [v for v in answers.values() if v is not None]
total_answered = len(answered)
score = sum(answered)

col1, col2 = st.columns([3, 1])
with col1:
    st.markdown(f"**Progress:** {total_answered} / {total_questions} questions answered")
    st.progress(total_answered / total_questions)
with col2:
    if total_answered > 0:
        st.metric("Score so far", f"{score}")

# ── Results ──────────────────────────────────────────────────────────────────
if total_answered == total_questions:
    if score >= 70:
        css_class, verdict = "strong", "✅ Strong Chance"
        description = "This opportunity is worth a full pursuit. Prioritise resources and go all in."
        color = "#28a745"
    elif score >= 45:
        css_class, verdict = "possible", "⚠️ Possible"
        description = "Proceed with caution. Identify your gaps and address them before committing fully."
        color = "#856404"
    else:
        css_class, verdict = "longshot", "❌ Long Shot"
        description = "Significant barriers exist. Consider whether the investment is justified."
        color = "#721c24"

    st.markdown(f"""
    <div class="verdict-box {css_class}">
        <div class="score-big" style="color:{color}">{score}</div>
        <div style="color:#777;font-size:14px">out of 100 points</div>
        <div class="verdict-label" style="color:{color}">{verdict}</div>
        <div style="margin-top:10px;font-size:15px;color:#333">{description}</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### Score Breakdown by Category")
    for category, questions in QUESTIONS.items():
        q_ids = [q["id"] for q in questions]
        cat_score = sum(answers[qid] for qid in q_ids if answers.get(qid) is not None)
        cat_max = MAX_PER_CATEGORY[category]
        st.markdown(f"**{category}** — {cat_score} / {cat_max}")
        st.progress(cat_score / cat_max)

    # ── Excel export ─────────────────────────────────────────────────────────
    def build_excel(answers, score, verdict, questions_map, max_map):
        wb = Workbook()
        ws = wb.active
        ws.title = "Scorecard"

        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 48
        ws.column_dimensions["D"].width = 3
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14

        # Helper styles
        def hdr_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Title block ──
        ws.merge_cells("B1:F1")
        title_cell = ws["B1"]
        title_cell.value = "🎯  Agency Opportunity Scorecard"
        title_cell.font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1a1a2e")
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 36

        # ── Score / Verdict summary ──
        ws.merge_cells("B2:C3")
        score_label = ws["B2"]
        score_label.value = f"Total Score:  {score} / 100"
        score_label.font = Font(name="Calibri", bold=True, size=14, color="1a1a2e")
        score_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        score_label.fill = PatternFill("solid", fgColor="F0F4FF")

        ws.merge_cells("E2:F3")
        verdict_clean = verdict.replace("✅ ", "").replace("⚠️ ", "").replace("❌ ", "")
        verdict_cell = ws["E2"]
        verdict_cell.value = verdict_clean
        if score >= 70:
            v_color, v_bg = "276221", "D4EDDA"
        elif score >= 45:
            v_color, v_bg = "856404", "FFF3CD"
        else:
            v_color, v_bg = "721c24", "F8D7DA"
        verdict_cell.font = Font(name="Calibri", bold=True, size=13, color=v_color)
        verdict_cell.fill = PatternFill("solid", fgColor=v_bg)
        verdict_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20

        # ── Legend row ──
        ws.row_dimensions[4].height = 8

        ws.merge_cells("B5:C5")
        ws["B5"].value = "Strong Chance (70+),  Possible (45–69),  Long Shot (0–44)"
        ws["B5"].font = Font(name="Calibri", italic=True, size=9, color="888888")
        ws["B5"].alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[5].height = 14

        # Spacer
        ws.row_dimensions[6].height = 6

        # ── Column headers ──
        row = 7
        for col, text in zip(["B", "C", "E", "F"], ["Criteria", "Notes / Rationale", "Max Pts", "Your Score"]):
            c = ws[f"{col}{row}"]
            c.value = text
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4361EE")
            c.alignment = Alignment(horizontal="center" if col in ("E", "F") else "left", vertical="center", indent=0 if col in ("E", "F") else 1)
            c.border = border
        ws.row_dimensions[row].height = 20

        # Category colour palette
        cat_colors = {
            "Competitive Landscape":  ("EEF0FF", "4361EE"),
            "Relationship & History": ("FFF0F5", "E63F7A"),
            "Brief & Process":        ("F0FFF4", "2E8B57"),
            "Commercial Opportunity": ("FFFAF0", "E07B39"),
            "Scope & Category":       ("F5F0FF", "7B2FBE"),
        }

        row = 8
        q_lookup = {q["id"]: q for qs in questions_map.values() for q in qs}

        for category, questions in questions_map.items():
            bg_light, bg_dark = cat_colors[category]
            cat_max = max_map[category]
            cat_score = sum(answers[q["id"]] for q in questions if answers.get(q["id"]) is not None)

            # Category header
            ws.merge_cells(f"B{row}:C{row}")
            cat_cell = ws[f"B{row}"]
            cat_cell.value = f"  {category}"
            cat_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cat_cell.fill = PatternFill("solid", fgColor=bg_dark)
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            cat_cell.border = border

            for col in ["E", "F"]:
                c = ws[f"{col}{row}"]
                c.fill = PatternFill("solid", fgColor=bg_dark)
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            ws[f"E{row}"].value = cat_max
            ws[f"F{row}"].value = cat_score
            ws.row_dimensions[row].height = 18
            row += 1

            # Question rows
            for q in questions:
                selected_pts = answers.get(q["id"])
                # Find selected label
                selected_label = ""
                for opt_label, opt_pts in q["options"]:
                    if opt_pts == selected_pts:
                        selected_label = opt_label
                        break

                b = ws[f"B{row}"]
                b.value = f"  {q['label']}"
                b.font = Font(name="Calibri", size=9, color="1a1a2e")
                b.fill = PatternFill("solid", fgColor=bg_light)
                b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                b.border = border

                c = ws[f"C{row}"]
                c.value = selected_label
                c.font = Font(name="Calibri", size=9, color="444444")
                c.fill = PatternFill("solid", fgColor=bg_light)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
                c.border = border

                e = ws[f"E{row}"]
                e.value = q["max"]
                e.font = Font(name="Calibri", size=9, color="888888")
                e.fill = PatternFill("solid", fgColor=bg_light)
                e.alignment = Alignment(horizontal="center", vertical="center")
                e.border = border

                f = ws[f"F{row}"]
                f.value = selected_pts if selected_pts is not None else ""
                f.font = Font(name="Calibri", bold=True, size=9,
                              color=bg_dark if selected_pts else "AAAAAA")
                f.fill = PatternFill("solid", fgColor=bg_light)
                f.alignment = Alignment(horizontal="center", vertical="center")
                f.border = border

                ws.row_dimensions[row].height = 16
                row += 1

            # Spacer
            ws.row_dimensions[row].height = 6
            row += 1

        # ── Grand total row ──
        ws.merge_cells(f"B{row}:C{row}")
        total_cell = ws[f"B{row}"]
        total_cell.value = "  TOTAL SCORE"
        total_cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        total_cell.fill = PatternFill("solid", fgColor="1a1a2e")
        total_cell.alignment = Alignment(horizontal="left", vertical="center")
        total_cell.border = border

        for col, val in [("E", 100), ("F", score)]:
            c = ws[f"{col}{row}"]
            c.value = val
            c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1a1a2e")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[row].height = 22

        # ── Sheet 2: options reference ──
        ws2 = wb.create_sheet("Scoring Guide")
        ws2.column_dimensions["A"].width = 3
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 48
        ws2.column_dimensions["D"].width = 10

        ws2.merge_cells("B1:D1")
        ws2["B1"].value = "Scoring Guide — All Options & Points"
        ws2["B1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        ws2["B1"].fill = PatternFill("solid", fgColor="1a1a2e")
        ws2["B1"].alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws2.row_dimensions[1].height = 28

        r = 3
        for category, questions in questions_map.items():
            bg_light, bg_dark = cat_colors[category]
            ws2.merge_cells(f"B{r}:D{r}")
            c = ws2[f"B{r}"]
            c.value = f"  {category}"
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=bg_dark)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws2.row_dimensions[r].height = 16
            r += 1

            for q in questions:
                ws2[f"B{r}"].value = q["id"]
                ws2[f"B{r}"].font = Font(name="Calibri", bold=True, size=9, color=bg_dark)
                ws2[f"B{r}"].fill = PatternFill("solid", fgColor=bg_light)
                ws2[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")
                ws2[f"B{r}"].border = border

                ws2[f"C{r}"].value = f"  {q['label']}"
                ws2[f"C{r}"].font = Font(name="Calibri", bold=True, size=9)
                ws2[f"C{r}"].fill = PatternFill("solid", fgColor=bg_light)
                ws2[f"C{r}"].alignment = Alignment(horizontal="left", vertical="center")
                ws2[f"C{r}"].border = border

                ws2[f"D{r}"].value = f"Max {q['max']}"
                ws2[f"D{r}"].font = Font(name="Calibri", size=9, color="888888")
                ws2[f"D{r}"].fill = PatternFill("solid", fgColor=bg_light)
                ws2[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")
                ws2[f"D{r}"].border = border
                ws2.row_dimensions[r].height = 15
                r += 1

                for opt_label, opt_pts in q["options"]:
                    ws2[f"C{r}"].value = f"      {opt_label}"
                    ws2[f"C{r}"].font = Font(name="Calibri", size=9, color="555555")
                    ws2[f"C{r}"].alignment = Alignment(horizontal="left", vertical="center")

                    ws2[f"D{r}"].value = opt_pts
                    ws2[f"D{r}"].font = Font(name="Calibri", size=9, color="1a1a2e")
                    ws2[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")
                    ws2.row_dimensions[r].height = 14
                    r += 1
                r += 1

        # Write to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    st.divider()
    st.markdown("### 📥 Download Your Results")

    col_l, col_r = st.columns([2, 1])
    with col_l:
        client_name = st.text_input("Client / Opportunity name (optional)", placeholder="e.g. Acme Co — Q2 2025 Pitch")
    with col_r:
        st.markdown("<br>", unsafe_allow_html=True)

    excel_buf = build_excel(answers, score, verdict, QUESTIONS, MAX_PER_CATEGORY)
    filename = f"Scorecard_{client_name.replace(' ', '_')}.xlsx" if client_name else "Agency_Opportunity_Scorecard.xlsx"

    st.download_button(
        label="⬇️  Download Scorecard (Excel)",
        data=excel_buf,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

elif total_answered > 0:
    st.info(f"Answer all {total_questions} questions to see your final verdict.")
else:
    st.info("Answer the questions above to score this opportunity.")
